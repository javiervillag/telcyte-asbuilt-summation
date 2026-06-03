from __future__ import annotations

import re
from pathlib import Path

CODE_PATTERN = re.compile(r"\b(UG|CD|MDU|COMP|FB|FX|PC|TL|CX|PT|SMC)-?(\d{1,3})(?!\.\d)\b", re.I)
NUMBER_PATTERN = r"(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?"
UNIT_PATTERN = r"(?:'|sq\.?\s*ft\.?|sqft)"
ZERO_PAD_EQUIVALENT_PREFIXES = {"UG", "CD", "MDU", "FB", "FX", "PC", "TL", "CX", "PT", "SMC"}
CodeKey = tuple[str, str]
TotalKey = tuple[CodeKey, str, str]


def code_key(code: str) -> CodeKey | None:
    match = CODE_PATTERN.search(code)
    if not match:
        return None
    prefix = match.group(1).upper()
    number = match.group(2)
    if prefix in ZERO_PAD_EQUIVALENT_PREFIXES:
        number = str(int(number))
    return (prefix, number)


def extract_codes_from_text(text: str) -> list[str]:
    codes: list[str] = []
    seen: set[CodeKey] = set()
    for match in CODE_PATTERN.finditer(text):
        raw = match.group(0)
        key = code_key(raw)
        if key and key not in seen:
            seen.add(key)
            codes.append(_format_code(match.group(1), match.group(2), raw))
    return codes


def total_line_key(line: str) -> TotalKey | None:
    code_match = CODE_PATTERN.search(line)
    if not code_match:
        return None
    key = code_key(code_match.group(0))
    if not key:
        return None
    remainder = line[code_match.end() :]
    qty_match = re.match(rf"\s*-\s*({NUMBER_PATTERN})(\s*{UNIT_PATTERN})?", remainder, re.I)
    if not qty_match:
        return None
    qty = _normalize_quantity(qty_match.group(1))
    unit = _normalize_unit(qty_match.group(2) or "")
    return (key, qty, unit)


def load_code_catalog(raw_codes: str = "", paths: str = "") -> dict[CodeKey, str]:
    catalog: dict[CodeKey, str] = {}
    for code in extract_codes_from_text(raw_codes):
        key = code_key(code)
        if key:
            catalog.setdefault(key, code)

    for raw_path in [p.strip() for p in paths.split(",") if p.strip()]:
        path = Path(raw_path)
        if not path.exists():
            continue
        for code in _codes_from_path(path):
            key = code_key(code)
            if key:
                catalog.setdefault(key, code)
    return catalog


def _codes_from_path(path: Path) -> list[str]:
    suffix = path.suffix.lower()
    if suffix in {".txt", ".csv", ".tsv"}:
        return extract_codes_from_text(path.read_text(errors="ignore"))
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            return []
        workbook = load_workbook(path, data_only=True, read_only=False)
        highlighted_parts: list[str] = []
        highlighted_sheet_parts: list[str] = []
        all_parts: list[str] = []
        try:
            for sheet in workbook.worksheets:
                sheet_is_highlighted = bool(sheet.sheet_properties.tabColor)
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        text = str(cell.value)
                        all_parts.append(text)
                        if _has_highlight_fill(cell.fill):
                            highlighted_parts.append(text)
                        if sheet_is_highlighted:
                            highlighted_sheet_parts.append(text)
        finally:
            workbook.close()
        if highlighted_parts:
            return extract_codes_from_text("\n".join(highlighted_parts))
        if highlighted_sheet_parts:
            return extract_codes_from_text("\n".join(highlighted_sheet_parts))
        return extract_codes_from_text("\n".join(all_parts))
    return extract_codes_from_text(path.read_text(errors="ignore"))


def _format_code(prefix: str, number: str, raw: str) -> str:
    raw = raw.strip()
    if "-" in raw:
        return raw
    return f"{prefix}-{number}"


def _normalize_quantity(value: str) -> str:
    number = float(value.replace(",", ""))
    return str(int(number)) if number.is_integer() else f"{number:g}"


def _normalize_unit(value: str) -> str:
    normalized = re.sub(r"[\s.]+", "", value.strip().lower())
    if normalized == "sqft":
        return "sqft"
    return value.strip().lower()


def _has_highlight_fill(fill) -> bool:
    if not fill or not fill.fill_type or fill.fill_type == "none":
        return False
    color = fill.fgColor
    if not color:
        return True
    if color.type == "rgb" and color.rgb in {"00000000", "00FFFFFF", "FFFFFFFF"}:
        return False
    if color.type == "indexed" and color.indexed in {64, 65}:
        return False
    return True
