from app.rate_cards import code_key, extract_codes_from_text, load_code_catalog, total_line_key


def test_code_key_treats_zero_padded_variants_as_same_code() -> None:
    assert code_key("UG-7") == ("UG", "7")
    assert code_key("UG-07") == ("UG", "7")
    assert code_key("PC01") == ("PC", "1")


def test_code_key_treats_two_digit_variants_as_same_for_supported_prefixes() -> None:
    for prefix in ["UG", "CD", "MDU", "FB", "FX", "PC", "TL", "CX", "PT", "SMC", "SME", "DP"]:
        assert code_key(f"{prefix}-7") == code_key(f"{prefix}-07")


def test_hyphenated_future_code_prefixes_are_detected_generically() -> None:
    assert code_key("DP-11") == ("DP", "11")
    assert code_key("SME-1") == ("SME", "1")
    assert extract_codes_from_text("DP-11 - 156' SME-01 - 1 ZZ-7 - 2") == ["DP-11", "SME-01", "ZZ-7"]


def test_unhyphenated_unknown_words_are_not_codes() -> None:
    assert extract_codes_from_text("MCA_RateCard_AM2 CX-05") == ["CX-05"]


def test_composite_codes_do_not_gain_zero_padding_equivalence() -> None:
    assert code_key("Comp-9") == ("COMP", "9")
    assert code_key("Comp-09") == ("COMP", "09")
    assert code_key("Comp-9") != code_key("Comp-09")


def test_eli_codes_are_ignored_for_asbuilt_totals() -> None:
    assert code_key("ELI-7") is None
    assert extract_codes_from_text("ELI-7 ELI-07 UG-7") == ["UG-7"]


def test_decimal_fragments_are_not_treated_as_codes() -> None:
    assert extract_codes_from_text("CX16.7 CX-05") == ["CX-05"]


def test_load_code_catalog_keeps_rate_card_display_code() -> None:
    catalog = load_code_catalog("UG-07, PC-01, Comp-13")
    assert catalog[("UG", "7")] == "UG-07"
    assert catalog[("PC", "1")] == "PC-01"
    assert catalog[("COMP", "13")] == "Comp-13"


def test_load_code_catalog_prefers_highlighted_xlsx_cells(tmp_path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Rate Card"
    sheet["A1"] = "UG-99"
    sheet["A2"] = "UG-07"
    sheet["A2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    path = tmp_path / "rate-card.xlsx"
    workbook.save(path)

    catalog = load_code_catalog(paths=str(path))

    assert ("UG", "7") in catalog
    assert ("UG", "99") not in catalog


def test_extract_codes_from_text_dedupes_variants() -> None:
    assert extract_codes_from_text("UG-7 UG-07 PC1 PC-01") == ["UG-7", "PC-1"]


def test_total_line_key_normalizes_supported_code_variants_and_spacing() -> None:
    assert total_line_key("CD-1 -1") == total_line_key("CD-01 - 1")
    assert total_line_key("MDU-5 - 2") == total_line_key("MDU-05 - 2")
    assert total_line_key("UG-7 - 10'") == total_line_key("UG-07 - 10'")
    assert total_line_key("UG-7 - 10'") == total_line_key("UG-7 - 10")
    assert total_line_key("UG-80 - 56sqft") == total_line_key("UG-80 - 56")


def test_total_line_key_keeps_composite_zero_padding_distinct() -> None:
    assert total_line_key("Comp-9 - 2") != total_line_key("Comp-09 - 2")
    assert total_line_key("Comp-9 - 2") == total_line_key("COMP-9 - 2")
